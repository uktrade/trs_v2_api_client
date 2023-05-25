import os
import shutil
import tempfile
import zipfile

import pikepdf
import pytest
from docx import Document
from lxml import etree
from openpyxl import Workbook, load_workbook

from v2_api_client.shared.upload_handler.metadata import Extractor


class TestDocumentMetadata:
    AUTHOR = "TRA"
    TITLE = "Extract Document Metadata"

    @pytest.fixture
    def document(self, request):
        file_types = {
            "pdf": self.create_pdf_document,
            "docx": self.create_docx_document,
            "xlsx": self.create_xlsx_document,
            "odf": self.create_odf_document,
        }

        for file_type in file_types:
            if file_type == request.param[0]:
                if len(request.param) > 1:
                    file_types[file_type](request.param[1])
                else:
                    file_types[file_type]()

                yield
                os.remove(getattr(self, f"{file_type}_file"))

    @pytest.mark.parametrize("document", [["pdf"]], indirect=True)
    def test_instantiate_document_metadata(self, document):
        assert isinstance(self.pdf_document, Extractor)

    @pytest.mark.parametrize("document", [["pdf"]], indirect=True)
    def test_extract_pdf_metadata(self, document):
        assert ["TRA"] == self.pdf.open_metadata()["dc:creator"]
        assert "Extract Document Metadata" == self.pdf.open_metadata()["dc:title"]

        with open(self.pdf_file, "rb") as file:
            raw_data = file.read()
            content_type = "application/pdf"

            sanitised_data = self.pdf_document(raw_data, content_type)
            metadata = str(pikepdf.open(sanitised_data).open_metadata())

            assert "TRA" not in metadata
            assert "Extract Document Metadata" not in metadata

    @pytest.mark.parametrize("document", [["docx"]], indirect=True)
    def test_extract_docx_metadata(self, document):
        assert self.docx.core_properties.author == "TRA"
        assert self.docx.core_properties.title == "Extract Document Metadata"

        with open(self.docx_file, "rb") as file:
            raw_data = file.read()
            content_type = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"

            sanitised_data = self.docx_document(raw_data, content_type)
            metadata = Document(sanitised_data).core_properties

            fields = [
                attr
                for attr in dir(metadata)
                if isinstance(getattr(metadata, attr), str) and not attr.startswith("_")
            ]

            for field in fields:
                assert "TRA" not in getattr(metadata, field)
                assert "Extract Document Metadata" not in getattr(metadata, field)

    @pytest.mark.parametrize("document", [["xlsx"]], indirect=True)
    def test_extract_xlsx_metadata(self, document):
        assert self.xlsx.properties.creator == "TRA"
        assert self.xlsx.properties.title == "Extract Document Metadata"

        with open(self.xlsx_file, "rb") as file:
            raw_data = file.read()
            content_type = (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            sanitised_data = self.xlsx_document(raw_data, content_type)
            metadata = load_workbook(sanitised_data).properties

            fields = [
                attr
                for attr in dir(metadata)
                if isinstance(getattr(metadata, attr), str)
                and not attr.startswith("_")
                and not attr == "tagname"
                and not attr == "namespace"
            ]

        for field in fields:
            assert "TRA" not in getattr(metadata, field)
            assert "Extract Document Metadata" not in getattr(metadata, field)

    @pytest.mark.parametrize(
        "document, content_type",
        [
            (["odf", "odt"], "application/vnd.oasis.opendocument.text"),
            (["odf", "ods"], "application/vnd.oasis.opendocument.spreadsheet"),
        ],
        indirect=["document"],
    )
    def test_extract_odf_metadata(self, document, content_type):
        with zipfile.ZipFile(self.odf_file, "r") as input_odf:
            for files in input_odf.infolist():
                if files.filename == "meta.xml":
                    root = etree.parse(input_odf.open("meta.xml")).getroot()

                    for fields in root[0]:
                        if fields.tag == "creator":
                            assert fields.text == "TRA"
                        if fields.tag == "title":
                            assert fields.text == "Extract Document Metadata"

        with open(self.odf_file, "rb") as file:
            raw_data = file.read()

            sanitised_data = self.odf_document(raw_data, content_type)

            with zipfile.ZipFile(sanitised_data, "r") as input_odf:
                for files in input_odf.infolist():
                    if files.filename == "meta.xml":
                        root = etree.parse(input_odf.open("meta.xml")).getroot()

                        for fields in root[0]:
                            if fields.tag == "creator":
                                assert fields.text != "TRA"
                            if fields.tag == "title":
                                assert fields.text != "Extract Document Metadata"

    def create_pdf_document(self):
        self.pdf = pikepdf.new()

        with self.pdf.open_metadata() as meta:
            meta["dc:creator"] = [self.AUTHOR]
            meta["dc:title"] = self.TITLE

        self.pdf_file = os.path.join(os.path.dirname(__file__), "fixture.pdf")
        self.pdf.save(self.pdf_file)

        self.pdf_document = Extractor()

    def create_docx_document(self):
        self.docx = Document()
        self.docx_metadata = self.docx.core_properties
        self.docx_metadata.author = self.AUTHOR
        self.docx_metadata.title = self.TITLE

        self.docx_file = os.path.join(os.path.dirname(__file__), "fixture.docx")
        self.docx.save(self.docx_file)

        self.docx_document = Extractor()

    def create_xlsx_document(self):
        self.xlsx = Workbook()
        self.xlsx_metadata = self.xlsx.properties
        self.xlsx_metadata.creator = self.AUTHOR
        self.xlsx_metadata.title = self.TITLE

        self.xlsx_file = os.path.join(os.path.dirname(__file__), "fixture.xlsx")
        self.xlsx.save(self.xlsx_file)

        self.xlsx_document = Extractor()

    def create_zip_file(self):
        self.create_pdf_document()
        self.create_xlsx_document()
        with tempfile.TemporaryDirectory() as tmpdirname:
            shutil.copy(self.pdf_file, os.path.join(tmpdirname, "fixture.pdf"))
            shutil.copy(self.xlsx_file, os.path.join(tmpdirname, "fixture.xlsx"))
            self.zip_file = shutil.make_archive(
                os.path.join(os.path.dirname(__file__), "fixture"), "zip", tmpdirname
            )

    def create_odf_document(self, file_type):
        odf_filepath = os.path.join(
            os.path.dirname(__file__), f"fixtures/sample.{file_type}"
        )
        temporary_file_descriptor, self.odf_file = tempfile.mkstemp(
            dir=os.path.dirname(odf_filepath)
        )
        os.close(temporary_file_descriptor)

        with zipfile.ZipFile(odf_filepath, "r") as input_odf:
            with zipfile.ZipFile(self.odf_file, "w") as output_odf:
                for file in input_odf.infolist():
                    if file.filename != "meta.xml":
                        output_odf.writestr(file, input_odf.read(file.filename))
                    else:
                        root = etree.parse(input_odf.open("meta.xml")).getroot()

                        creator = etree.SubElement(root[0], "creator")
                        title = etree.SubElement(root[0], "title")

                        creator.text = self.AUTHOR
                        title.text = self.TITLE

                        metadata = etree.tostring(root)

        with zipfile.ZipFile(self.odf_file, "a", zipfile.ZIP_DEFLATED) as zf:
            zf.writestr("meta.xml", metadata)

        self.odf_document = Extractor()
